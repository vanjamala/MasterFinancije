import streamlit as st
import pandas as pd
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

st.title("Prilagodba tablice MasterTeama i Konto - PRIJEVOZ")
st.write("Broj dana troška za prijevoz.")

#Enter month and year for the dates
# User input for month and year
month = st.text_input("Unesite mjesec (npr. '03'):", value='03')
year = st.text_input("Unesite godinu (npr. '2025'):", value='2025')

# Check if the month and year are entered correctly
if not month or not year:
    st.warning("Molimo unesite mjesec i godinu prije nego što nastavite.")
else:
    # Upload files
    uploaded_masterteam = st.file_uploader("Učitajte MasterTeam evidenciju", type=["xls", "xlsx"])
    uploaded_pn = st.file_uploader("Učitajte datoteku službenih putovanja", type=["xls", "xlsx"])

    # Process files if uploaded and inputs are valid
    if uploaded_masterteam and uploaded_pn and st.button('Spoji podatke i pripremi izvještaj'):
        df = pd.read_excel(uploaded_masterteam, header=3, engine='openpyxl')
        
        df_pn = pd.read_excel(uploaded_pn, header=3, engine='xlrd')
                # Filter out unnamed columns
        df = df.loc[:, ~df.columns.str.contains('^Unnamed', na=False)]
        # Drop rows where 'Fond' equals 'Fond' or is empty
        df = df[~df['Fond'].isin(['Fond']) & df['Fond'].notna()]
        # Convert values in the 'Prezime' column to strings and filter rows
        df = df[df['Prezime'].astype(str).apply(lambda x: not x.isnumeric())]
        # Handling missing values and concatenating Prezime and Ime columns
        df['PREZIME i IME'] = df['Prezime'].fillna('') + ' ' + df['Ime'].fillna('')
        # Rearrange columns: make PREZIME i IME the first column
        columns_order = ['PREZIME i IME'] + [col for col in df.columns if col not in ['PREZIME i IME', 'Prezime', 'Ime']]
        df = df[columns_order]
        # Filter the DataFrame for "Jovanović Rimac Ivana" in uppercase
        filtered_row = df[df["PREZIME i IME"].str.upper() == "JOVANOVIĆ RIMAC IVANA"]

        # Select the first 5 columns and print them
        print(filtered_row.iloc[:, :5])

        # Step 1: Remove columns that start with "Su" or "Ne" and rename totals to hours colums
        columns_to_remove = [col for col in df.columns if col.startswith("Su") or col.startswith("Ne")]
        df.drop(columns=columns_to_remove, inplace=True)

        #Rename hour columns
        # List of columns to rename
        columns_to_rename = ['Rad', 'Rad od kuće','Praznik', 'G.O.', 'Dopust', 'Bolo.','HZZO','Školovanje','Sl. dan']

        # Rename the specified columns by adding 'sati' to each column name
        df.rename(columns={col: col + ' sati' for col in columns_to_rename}, inplace=True)

        # Step 2: Extract day from column headers and create the new date (MM.DD.YYYY)
        new_column_headers = []
        for col in df.columns:
            # Clean the column name by removing any unwanted characters (e.g., line breaks)
            clean_col = str(col).strip().replace("\n", "")  # Remove line breaks

            # Attempt to find a day number (it can be one or two digits)
            day_match = re.search(r'(\d{1,2})', clean_col)  # Match one or two digits

            #print(day_match)

            if day_match:
                # Get the day as a two-digit number (e.g., 1 becomes 01)
                day = day_match.group(1).zfill(2)
                # Create the new date format: MM.DD.YYYY
                new_date = f"{day}.{month}.{year}"
                new_column_headers.append(new_date)
            else:
                # If no day number in the header, just use the original header
                new_column_headers.append(col)
        print(new_column_headers)
        # Step 3: Update the DataFrame column names with the new date format
        df.columns = new_column_headers

        # Step 4: Replace numeric values in date columns with "RR"
        # Identify the columns that are in date format (MM.DD.YYYY)
        for col in df.columns:
            if re.match(r'\d{2}\.\d{2}\.\d{4}', col):  # Match MM.DD.YYYY format
                df[col] = df[col].apply(lambda x: "RR" if (not pd.isna(x)) and (isinstance(x, (int, float)) or str(x).replace('.', '', 1).isdigit()) else x)

        # Step 5: Import putne naloge,transform data and update df with službena putovanja
        df_pn = df_pn[df_pn['Broj PN\n'] != 'SVEUKUPNO']

        # Convert 'Dat. Polaska' and 'Dat. Povratka' to datetime format (if they are not already in datetime format)
        df_pn["Dat. Polaska"] = pd.to_datetime(df_pn["Dat. Polaska"], errors='coerce')
        df_pn["Dat. Povratka"] = pd.to_datetime(df_pn["Dat. Povratka"], errors='coerce')

        # Expand each row into multiple rows for each day in the GO period
        expanded_rows = []
        for _, row in df_pn.iterrows():
            # Generate all dates between 'Dat. Polaska' and 'Dat. Povratka'
            if pd.notna(row["Dat. Polaska"]) and pd.notna(row["Dat. Povratka"]):
                date_range = pd.date_range(row["Dat. Polaska"], row["Dat. Povratka"])  # Generate all dates
                for date in date_range:
                    expanded_rows.append({
                        "Prezime Ime": row['Prezime i ime'],  # Use the correct column name
                        "Datum": date.strftime("%d.%m.%Y"),
                        "Razlog odsutnosti": row["Zadatak službenog puta"]
                    })

        # Convert the list of rows into a DataFrame
        df_expanded = pd.DataFrame(expanded_rows)

        # Loop through each row in df_expanded to update the df table
        for _, row in df_expanded.iterrows():
            person = row['Prezime Ime']
            date_str = str(row['Datum'])  # Make sure date is a string (DD.MM.YYYY)

            # Debug: Check what date and person we are processing
            #print(f"Checking for person: {person}, Date: {date_str}")

            # Check if the date exists as a column in df
            if date_str in df.columns:
                # Debug: If a match is found, print the action
                #print(f"Updating {person} on {date_str} to 'SP'")

                # Update the corresponding cell for the person and date to "SP"
                df.loc[df['PREZIME i IME'] == person, date_str] = "SP"
            #else:
                # Debug: If no match found, print a message
                #print(f"No matching column for date {date_str}")

        # Step 6: Insert new columns RR, GO, DODATI, and UKUPNO after the date columns
        # Identify the columns that are in date format (MM.DD.YYYY)
        date_columns = [col for col in df.columns if re.match(r'\d{2}\.\d{2}\.\d{4}', col)]

        # For each row, calculate RR, GO, DODATI, and UKUPNO
        df['RR dani'] = df[date_columns].apply(lambda row: row.astype(str).tolist().count('RR'), axis=1)
        df['GO dani'] = df[date_columns].apply(lambda row: row.astype(str).tolist().count('G'), axis=1)
        def count_GO_sa(row, date_columns):
            count = 0
            # Only loop through date columns (from the 3rd column onward)
            for i in range(2, len(date_columns) + 2):  # Adjusted to start at index 2
                value = row[i]

                if value != 'G':
                    continue  # Skip if it's "RR"

                # Check the consecutive sequence before and after the current value
                left_count = 0
                right_count = 0

                # Check to the left (before the current value)
                for left in range(i - 1, 2 - 1, -1):  # Loop backwards, but only up to date_columns
                    if row[left] != 'RR' and row[left] != 'SP':  # Non-"RR" and non "SP" value
                        left_count += 1
                    else:
                        break

                # Check to the right (after the current value)
                for right in range(i + 1, len(date_columns) + 2):  # Loop forwards, but only up to date_columns
                    if row[right] != 'RR' and row[right] != 'SP':  # Non-"RR" and non "SP" value
                        right_count += 1
                    else:
                        break

                # If the sum of left_count and right_count is 2 or more, we have a streak
                if left_count + right_count >= 2:
                    continue  # Don't count this value, as it's part of a streak

                # Otherwise, count this value
                count += 1

            return count
        df['GO s prijevozom'] = df.apply(lambda row: count_GO_sa(row, date_columns), axis=1)
        df['GO bez prijevoza'] = df['GO dani'] -df['GO s prijevozom']
        def SP_count(row, date_columns):
            count = 0
            # Loop through the date columns (starting from the 3rd column)
            for i in range(2, len(date_columns) + 2):  # Adjusted to start at index 2 (3rd column)
                value = row[i]

                if value != 'SP':  # We're only interested in 'SP'
                    continue

                # Check consecutive sequence before and after the current "SP" value
                left_count = 0
                right_count = 0

                # Check to the left (before the current "SP")
                for left in range(i - 1, 2 - 1, -1):  # Loop backwards, but only up to the first date column
                    if row[left] == 'SP':  # Non-"SP" value
                        left_count += 1
                    else:
                        break

                # Check to the right (after the current "SP")
                for right in range(i + 1, len(date_columns) +2 ):  # Loop forwards, but only up to the last date column
                    if row[right] == 'SP':  # Non-"SP" value
                        right_count += 1
                    else:
                        break

                # If the sum of left_count and right_count is 2 or more, we have a streak
                if left_count >=1 and right_count >= 1 or left_count>1:
                    continue  # Don't count this "SP", it's part of a streak


                # If the sum of left_count and right_count is 2 or more, we have a streak
                #if left_count + right_count >= 2:
                    #continue  # Don't count this "SP", it's part of a streak

                # Otherwise, count this "SP"
                count += 1
            return count
        df['SP prijevoz'] = df.apply(lambda row: SP_count(row, date_columns), axis=1)
        df['SP dani'] = df[date_columns].apply(lambda row: row.astype(str).tolist().count('SP'), axis=1)
        def count_non_rr(row, date_columns):
            count = 0
            # Only loop through date columns (from the 3rd column onward)
            for i in range(2, len(date_columns) + 2):  # Adjusted to start at index 2
                value = row[i]

                if value == 'RR' or value == 'SP':
                    continue  # Skip if it's "RR"

                # Check the consecutive sequence before and after the current value
                left_count = 0
                right_count = 0

                # Check to the left (before the current value)
                for left in range(i - 1, 2 - 1, -1):  # Loop backwards, but only up to date_columns
                    if row[left] != 'RR' and row[left] != 'SP':  # Non-"RR" and non "SP" value
                        left_count += 1
                    else:
                        break

                # Check to the right (after the current value)
                for right in range(i + 1, len(date_columns) + 2):  # Loop forwards, but only up to date_columns
                    if row[right] != 'RR' and row[right] != 'SP':  # Non-"RR" and non "SP" value
                        right_count += 1
                    else:
                        break

                # If the sum of left_count and right_count is 2 or more, we have a streak
                if left_count + right_count >= 2:
                    continue  # Don't count this value, as it's part of a streak

                # Otherwise, count this value
                count += 1

            return count

        # Apply the function to the rows of the date columns in df
        df['DODATI dani'] = df.apply(lambda row: count_non_rr(row, date_columns), axis=1)+df['SP prijevoz']
        df['UKUPNO dani'] = df['RR dani'] + df['DODATI dani']

        # Step 6: Insert the new columns after the date columns (displace the previous ones)
        date_column_index = 2 + len(date_columns)  # The number of date columns determines the insertion point

        # Insert the new columns at the right place
        df.insert(date_column_index, 'RR dani', df.pop('RR dani'))
        df.insert(date_column_index + 1, 'GO dani', df.pop('GO dani'))
        df.insert(date_column_index + 2, 'GO s prijevozom', df.pop('GO s prijevozom'))
        df.insert(date_column_index + 3, 'GO bez prijevoza', df.pop('GO bez prijevoza'))
        df.insert(date_column_index + 4, 'SP dani', df.pop('SP dani'))
        df.insert(date_column_index + 5, 'SP prijevoz', df.pop('SP prijevoz'))
        df.insert(date_column_index + 6, 'DODATI dani', df.pop('DODATI dani'))
        df.insert(date_column_index + 7, 'UKUPNO dani', df.pop('UKUPNO dani'))

        # Step 7: Insert SP sati
        # First, 'SP sati' column by multiplying 'SP dani' by 8
        df['SP sati'] = df['SP dani'] * 8

        # Insert 'SP sati' after the 'Dopust' column
        dopust_index = df.columns.get_loc('Dopust sati')  # Get the index of the 'Dopust' column
        df.insert(dopust_index + 1, 'SP sati', df.pop('SP sati'))  # Insert 'SP sati' after 'Dopust'

        #Step 8: Insert ukupno sati
        # Assuming the columns 'Praznik', 'G.O.', '#G.O.', 'Dopust', 'HZZO' exist in df
        df['Ukupno sati bez sati SP (SP sati uračunato u Rad sati)'] = df[['Rad sati', 'Rad od kuće sati', 'Praznik sati', 'G.O. sati', 'Dopust sati', 'Bolo. sati','HZZO sati','Školovanje sati','Sl. dan sati']].sum(axis=1)
        # Report
        st.write(df)
        # Create an Excel workbook and add the DataFrame to it
        wb = Workbook()
        sheet = wb.active

        # Add DataFrame to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Define pastel colors
        green_fill = PatternFill(start_color='C2EABD', end_color='C2EABD', fill_type='solid')  # Pastel green
        yellow_fill = PatternFill(start_color='FFF4C1', end_color='FFF4C1', fill_type='solid')  # Pastel yellow
        red_fill = PatternFill(start_color='F7C6C7', end_color='F7C6C7', fill_type='solid')  # Pastel red

        # Apply colors based on "Status" column
        for row in sheet.iter_rows(min_row=2):  # Assuming first row is a header  # Focus on first column
            for cell in row:
                #print(f"Cell {cell.coordinate} has value: {cell.value}")  # Debug to confirm values
                if cell.value == "RR":
                    cell.fill = green_fill
                    #print(f"Cell {cell.coordinate} formatted with green_fill")  # Debug
                elif cell.value == "G":
                    cell.fill = yellow_fill
                elif cell.value == "SP":
                    cell.fill = red_fill

        # Save the workbook to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reset pointer to start

        # Create a flag to track when the download is complete
        if "downloaded" not in st.session_state:
            st.session_state["downloaded"] = False

        # Generate and offer the file for download
        st.download_button(
            label="Preuzmi tablicu u boji",
            data=output,
            file_name="colored_output.xlsx",
            mime="application/vnd.ms-excel",
            on_click=lambda: st.session_state.update({"downloaded": True})  # Set the flag
        )

        # Clear the uploaded files **only after the user has clicked download**
        if st.session_state["downloaded"]:
            st.session_state.pop("uploaded_masterteam", None)
            st.session_state.pop("uploaded_pn", None)
            st.session_state["upload_reset"] = st.session_state.get("upload_reset", 0) + 1
            st.session_state["downloaded"] = False  # Reset the flag
            st.experimental_rerun()  # Rerun to refresh the file uploader


       

        
